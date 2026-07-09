#!/usr/bin/env python3
"""列出部门周报 workbook 全部子表，并标出 config 会选中的目标子表。"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "lib"))

from sheet_utils import resolve_dept_sheet


def load_config(path: Path) -> dict:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def read_workbook_sheet_names(path: Path) -> tuple[list[str], object | None]:
    """返回 (子表名列表, 可选 openpyxl workbook)。"""
    if path.suffix.lower() == ".ksheet":
        from sheet_utils import list_ksheet_sheet_names

        return list_ksheet_sheet_names(path), None

    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit("请安装 openpyxl: pip install openpyxl") from e

    wb = load_workbook(path, read_only=True, data_only=True)
    return list(wb.sheetnames), wb


def main() -> int:
    parser = argparse.ArgumentParser(description="列出部门周报表格子表")
    parser.add_argument("--config", type=Path, default=Path("config.json"))
    parser.add_argument(
        "--input",
        type=Path,
        required=True,
        help="部门周报本地路径（.ksheet / .xlsx / .et）",
    )
    args = parser.parse_args()

    cfg = load_config(args.config)
    names, wb = read_workbook_sheet_names(args.input)
    resolved, reason = resolve_dept_sheet(names, cfg)

    sheets = []
    if wb is not None:
        for sn in names:
            ws = wb[sn]
            sheets.append({"name": sn, "max_row": ws.max_row, "max_column": ws.max_column})
        wb.close()
    else:
        for sn in names:
            sheets.append({"name": sn})

    payload = {
        "workbook": str(args.input),
        "format": args.input.suffix.lower().lstrip(".") or "unknown",
        "sheet_count": len(names),
        "sheets": sheets,
        "all_sheets": names,
        "resolved_sheet": resolved,
        "resolve_reason": reason,
        "team_name": cfg.get("team_name"),
        "fourth_dept_name": (cfg.get("dept_sheet") or {}).get("fourth_dept_name"),
        "configured_sheet_name": (cfg.get("dept_sheet") or {}).get("sheet_name"),
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0 if resolved else 1


if __name__ == "__main__":
    raise SystemExit(main())
