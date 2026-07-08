#!/usr/bin/env python3
"""根据部门周报与 extracted.json 生成 Excel 补丁计划（支持多子表）。"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "lib"))

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore

from sheet_utils import (
    find_cell_in_rows,
    find_sheet_in_kdc,
    index_to_col,
    kdc_sheet_to_rows,
    list_sheet_names_from_markdown,
    parse_sheet_table_from_markdown,
    resolve_dept_sheet,
    worksheet_to_rows,
)


def load_json(path: Path) -> dict:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def parse_md_tables(md: str) -> list[list[list[str]]]:
    tables: list[list[list[str]]] = []
    current: list[list[str]] = []
    for line in md.splitlines():
        if line.strip().startswith("|"):
            cells = [c.strip() for c in line.strip("|").split("|")]
            if all(set(c) <= set("-: ") for c in cells):
                continue
            current.append(cells)
        else:
            if current:
                tables.append(current)
                current = []
    if current:
        tables.append(current)
    return tables


def plan_from_xlsx(cfg: dict, extracted: dict, xlsx_path: Path) -> dict:
    if load_workbook is None:
        raise SystemExit("请安装 openpyxl: pip install openpyxl")

    wb = load_workbook(xlsx_path, data_only=True)
    sheet_names = list(wb.sheetnames)
    resolved, reason = resolve_dept_sheet(sheet_names, cfg)
    if not resolved:
        return {"week": cfg.get("week"), "resolved_sheet": None, "resolve_reason": reason, "patches": []}

    ws = wb[resolved]
    rows = worksheet_to_rows(ws)
    by_name = {m["name"]: m["content"] for m in extracted.get("members", [])}
    plan = []

    for member in cfg.get("members", []):
        target = dict(member.get("target", {}))
        if target.get("type") != "sheet_cell":
            continue
        target["sheet"] = resolved
        pos = find_cell_in_rows(rows, member, target, cfg)
        entry = {
            "name": member["name"],
            "content": by_name.get(member["name"], ""),
            "sheet": resolved,
            "target": target,
        }
        if pos:
            row_idx, col_idx = pos
            entry["row"] = row_idx + 1
            entry["col"] = col_idx + 1
            entry["cell"] = f"{index_to_col(col_idx)}{row_idx + 1}"
            entry["status"] = "ready"
        else:
            entry["status"] = "not_found"
        plan.append(entry)

    return {
        "week": cfg.get("week"),
        "resolved_sheet": resolved,
        "resolve_reason": reason,
        "patches": plan,
    }


def plan_from_kdc_json(cfg: dict, extracted: dict, kdc_path: Path) -> dict:
    data = json.loads(kdc_path.read_text(encoding="utf-8"))
    raw = data.get("raw") or data
    sheet_names = [s.get("name", "") for s in (raw.get("doc") or raw).get("sheets", [])]
    resolved, reason = resolve_dept_sheet([n for n in sheet_names if n], cfg)
    if not resolved:
        return {"week": cfg.get("week"), "resolved_sheet": None, "resolve_reason": reason, "patches": []}

    sheet = find_sheet_in_kdc(raw, resolved)
    if not sheet:
        return {
            "week": cfg.get("week"),
            "resolved_sheet": resolved,
            "resolve_reason": reason,
            "patches": [],
            "error": f"KDC 中未找到子表: {resolved}",
        }

    rows = kdc_sheet_to_rows(sheet)
    by_name = {m["name"]: m["content"] for m in extracted.get("members", [])}
    plan = []

    for member in cfg.get("members", []):
        target = dict(member.get("target", {}))
        if target.get("type") != "sheet_cell":
            continue
        target["sheet"] = resolved
        pos = find_cell_in_rows(rows, member, target, cfg)
        entry = {
            "name": member["name"],
            "content": by_name.get(member["name"], ""),
            "sheet": resolved,
            "target": target,
        }
        if pos:
            row_idx, col_idx = pos
            entry["row"] = row_idx + 1
            entry["col"] = col_idx + 1
            entry["cell"] = f"{index_to_col(col_idx)}{row_idx + 1}"
            entry["status"] = "ready"
        else:
            entry["status"] = "not_found"
        plan.append(entry)

    return {
        "week": cfg.get("week"),
        "resolved_sheet": resolved,
        "resolve_reason": reason,
        "patches": plan,
        "source": "kdc_json",
    }


def plan_from_markdown(cfg: dict, extracted: dict, md_path: Path, table_index: int) -> dict:
    md = md_path.read_text(encoding="utf-8")
    sheet_names = list_sheet_names_from_markdown(md)
    resolved, reason = resolve_dept_sheet(sheet_names, cfg)
    if not resolved:
        return {"week": cfg.get("week"), "resolved_sheet": None, "resolve_reason": reason, "patches": []}

    table = parse_sheet_table_from_markdown(md, resolved)
    if not table:
        tables = parse_md_tables(md)
        if not tables:
            raise SystemExit("部门 Markdown 中未解析到表格")
        table = tables[table_index]

    by_name = {m["name"]: m["content"] for m in extracted.get("members", [])}
    plan = []

    for member in cfg.get("members", []):
        target = dict(member.get("target", {}))
        if target.get("type") != "sheet_cell":
            continue
        target["sheet"] = resolved
        pos = find_cell_in_rows(table, member, target, cfg)
        entry = {
            "name": member["name"],
            "content": by_name.get(member["name"], ""),
            "sheet": resolved,
            "target": target,
        }
        if pos:
            row_idx, col_idx = pos
            entry["row"] = row_idx + 1
            entry["col"] = col_idx + 1
            entry["cell"] = f"{index_to_col(col_idx)}{row_idx + 1}"
            entry["status"] = "ready"
        else:
            entry["status"] = "not_found"
        plan.append(entry)

    return {"week": cfg.get("week"), "resolved_sheet": resolved, "resolve_reason": reason, "patches": plan}


def main() -> int:
    parser = argparse.ArgumentParser(description="生成部门表格补丁计划")
    parser.add_argument("--config", type=Path, default=Path("config.json"))
    parser.add_argument("--extracted", type=Path, required=True)
    parser.add_argument("--input-xlsx", type=Path, help="部门周报 xlsx（推荐，支持多子表）")
    parser.add_argument("--dept-kdc-json", type=Path, help="部门周报 KDC JSON（推荐，ksheet 多行单元格）")
    parser.add_argument("--dept-markdown", type=Path, help="部门周报 Markdown（备选）")
    parser.add_argument("--output", type=Path, default=Path(".cache/patch-plan.json"))
    parser.add_argument("--table-index", type=int, default=0)
    args = parser.parse_args()

    cfg = load_json(args.config)
    extracted = load_json(args.extracted)

    if args.input_xlsx:
        payload = plan_from_xlsx(cfg, extracted, args.input_xlsx)
    elif args.dept_kdc_json:
        payload = plan_from_kdc_json(cfg, extracted, args.dept_kdc_json)
    elif args.dept_markdown:
        payload = plan_from_markdown(cfg, extracted, args.dept_markdown, args.table_index)
    else:
        print("请指定 --input-xlsx、--dept-kdc-json 或 --dept-markdown", file=sys.stderr)
        return 2

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(payload, ensure_ascii=False, indent=2))

    if not payload.get("resolved_sheet"):
        return 1
    missing = [p["name"] for p in payload.get("patches", []) if p.get("status") != "ready"]
    return 1 if missing else 0


if __name__ == "__main__":
    raise SystemExit(main())
