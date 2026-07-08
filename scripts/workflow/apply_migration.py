#!/usr/bin/env python3
"""一键写回：局部 patch ksheet → 校验 → wps365-read update 上传。"""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "lib"))
from paths import CACHE, SCRIPTS_ROOT, SKILL_ROOT, setup_sys_path  # noqa: E402

setup_sys_path("patch")
from wps365_bridge import find_wps365_read_root, resolve_wps_sid, run_drive  # noqa: E402

EXTRACT = SCRIPTS_ROOT / "extract"
PLAN = SCRIPTS_ROOT / "plan"
PATCH = SCRIPTS_ROOT / "patch"
WORKFLOW = SCRIPTS_ROOT / "workflow"


def load_json_file(path: Path) -> dict:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def find_dept_ksheet(cache: Path, cfg: dict | None = None) -> Path | None:
    exclude = {"dept-report-patched.ksheet", "test-patch.ksheet", "dept-report-patched-zip.ksheet"}
    candidates = [
        p
        for p in cache.glob("*.ksheet")
        if p.name not in exclude and not p.name.startswith("test")
    ]
    title = ((cfg or {}).get("dept_report") or {}).get("title", "")
    titled = [p for p in candidates if title and title in p.name]
    if titled:
        candidates = titled

    if not candidates:
        return None

    def quality(path: Path) -> tuple[int, int]:
        try:
            import zipfile

            with zipfile.ZipFile(path) as z:
                if "customXml/item2.xml" not in z.namelist():
                    return (0, path.stat().st_size)
                links = z.read("customXml/item2.xml").count(b"hypersublink")
                return (links, path.stat().st_size)
        except OSError:
            return (0, 0)

    return max(candidates, key=quality)


def restore_raw_content(extracted_path: Path) -> None:
    data = json.loads(extracted_path.read_text(encoding="utf-8"))
    for m in data.get("members", []):
        if m.get("content_raw"):
            m["content"] = m["content_raw"]
            m["char_count"] = len(m["content"])
    extracted_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def run_step(cmd: list[str]) -> int:
    print(f">>> {' '.join(cmd)}", flush=True)
    return subprocess.call(cmd, cwd=str(SKILL_ROOT))


def verify_patched_cells(patched: Path, plan: dict, extracted: dict) -> list[dict]:
    by_name = {m["name"]: m["content"] for m in extracted.get("members", [])}
    errors: list[dict] = []

    if patched.suffix.lower() == ".ksheet":
        import zipfile

        from patch_ksheet_zip import read_cell_text, resolve_sheet_path

        with zipfile.ZipFile(patched) as z:
            sheet_paths: dict[str, str] = {}
            for patch in plan.get("patches", []):
                if patch.get("status") != "ready":
                    continue
                name = patch["name"]
                expected = by_name.get(name, "")
                sheet = patch.get("sheet") or patch["target"]["sheet"]
                if sheet not in sheet_paths:
                    sp = resolve_sheet_path(z, sheet)
                    if not sp:
                        errors.append({"name": name, "reason": f"sheet 不存在: {sheet}"})
                        continue
                    sheet_paths[sheet] = sp
                cell_ref = patch.get("cell") or f"C{patch['row']}"
                actual = read_cell_text(patched, sheet_paths[sheet], cell_ref)
                if actual != expected:
                    errors.append(
                        {
                            "name": name,
                            "cell": patch.get("cell"),
                            "expected_len": len(expected),
                            "actual_len": len(str(actual or "")),
                            "match": actual == expected,
                        }
                    )
        return errors

    import shutil
    import tempfile
    from openpyxl import load_workbook

    tmp = Path(tempfile.mkdtemp(prefix="verify-"))
    try:
        xlsx = tmp / "v.xlsx"
        shutil.copy2(patched, xlsx)
        wb = load_workbook(xlsx, data_only=True)
        for patch in plan.get("patches", []):
            if patch.get("status") != "ready":
                continue
            name = patch["name"]
            expected = by_name.get(name, "")
            sheet = patch.get("sheet") or patch["target"]["sheet"]
            ws = wb[sheet]
            actual = ws.cell(row=patch["row"], column=patch["col"]).value
            if actual != expected:
                errors.append(
                    {
                        "name": name,
                        "cell": patch.get("cell"),
                        "expected_len": len(expected),
                        "actual_len": len(str(actual or "")),
                        "match": actual == expected,
                    }
                )
        return errors
    finally:
        import shutil as sh

        sh.rmtree(tmp, ignore_errors=True)


def main() -> int:
    parser = argparse.ArgumentParser(description="周报迁移写回部门 ksheet")
    parser.add_argument("--config", default="config.json")
    parser.add_argument("--input", type=Path, help="部门 ksheet 本地路径，默认 .cache 中最新下载")
    parser.add_argument("--upload", action="store_true", help="校验通过后上传到云文档")
    parser.add_argument("--skip-bullet-format", action="store_true", help="保留 otl 原文 - 列表符（默认会转为 •/◦）")
    parser.add_argument("--skip-download", action="store_true", help="上传前不重新下载云文档")
    parser.add_argument("--link-id", default=None, help="部门文档 link_id，默认读 config")
    parser.add_argument(
        "--keep-cache",
        action="store_true",
        help="写回成功后保留 .cache 中间产物（默认会清理）",
    )
    args = parser.parse_args()

    cfg = load_json_file(SKILL_ROOT / args.config)

    restore_raw_content(CACHE / "extracted.json")

    if not args.skip_bullet_format:
        if run_step(
            [
                sys.executable,
                str(EXTRACT / "format_otl_for_ksheet.py"),
                "--input",
                str(CACHE / "extracted.json"),
                "--kdc-json",
                str(CACHE / "dept-content.json"),
                "--sheet-name",
                (cfg.get("dept_sheet") or {}).get("sheet_name") or "",
                "--in-place",
            ]
        ) != 0:
            return 1

    link_id = args.link_id or (cfg.get("dept_report") or {}).get("link_id")
    if args.upload and not args.skip_download and link_id:
        sid, _ = resolve_wps_sid(cfg)
        wps_root = find_wps365_read_root(cfg)
        if sid and wps_root:
            print(f">>> download {link_id}", flush=True)
            run_drive(wps_root, ["download", link_id, "--dir", str(CACHE)])

    input_path = args.input or find_dept_ksheet(CACHE, cfg)
    if not input_path or not input_path.exists():
        print("未找到部门 ksheet，请先运行 scripts/workflow/preflight.py", file=sys.stderr)
        return 1

    week = cfg.get("week")
    if week and input_path.suffix.lower() == ".ksheet":
        ensure_cmd = [
            sys.executable,
            str(WORKFLOW / "ensure_week_column.py"),
            "--config",
            args.config,
        ]
        if run_step(ensure_cmd) != 0:
            return 1
        input_path = find_dept_ksheet(CACHE, cfg) or input_path

    plan_src = ["--dept-ksheet", str(input_path)] if input_path.suffix.lower() == ".ksheet" else ["--dept-kdc-json", str(CACHE / "dept-content.json")]
    if run_step(
        [
            sys.executable,
            str(PLAN / "plan_sheet_patches.py"),
            "--config",
            args.config,
            *plan_src,
            "--extracted",
            str(CACHE / "extracted.json"),
            "--output",
            str(CACHE / "patch-plan.json"),
        ]
    ) != 0:
        return 1

    plan = load_json_file(CACHE / "patch-plan.json")
    extracted = load_json_file(CACHE / "extracted.json")

    ready = [p for p in plan.get("patches", []) if p.get("status") == "ready"]
    if len(ready) != len(cfg.get("members", [])):
        missing = [p["name"] for p in plan.get("patches", []) if p.get("status") != "ready"]
        print(f"补丁计划未全部 ready，缺失: {missing}", file=sys.stderr)
        return 1

    output = CACHE / "dept-report-patched.ksheet"
    patch_script = PATCH / "patch_ksheet_zip.py"
    if input_path.suffix.lower() != ".ksheet":
        patch_script = PATCH / "patch_sheet.py"
    patch_cmd = [
        sys.executable,
        str(patch_script),
        "--config",
        args.config,
        "--input",
        str(input_path),
        "--plan",
        str(CACHE / "patch-plan.json"),
        "--extracted",
        str(CACHE / "extracted.json"),
        "--output",
        str(output),
    ]
    print(f">>> {' '.join(patch_cmd)}", flush=True)
    if subprocess.call(patch_cmd, cwd=str(SKILL_ROOT)) != 0:
        return 2

    verify_errors = verify_patched_cells(output, plan, extracted)
    if verify_errors:
        print(json.dumps({"error": "写后校验失败", "details": verify_errors}, ensure_ascii=False, indent=2), file=sys.stderr)
        return 3

    if output.suffix.lower() == ".ksheet":
        from patch_ksheet_zip import count_hypersublinks

        link_count = count_hypersublinks(output)
        if link_count == 0:
            print(
                json.dumps(
                    {
                        "error": "写后 customXml 超链接为空，请使用含 hypersublink 的原始 .ksheet 下载件",
                        "hint": "勿用 openpyxl save；优先 .cache 中带 hypersublink 的 ksheet 文件",
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                file=sys.stderr,
            )
            return 3

    report = {
        "status": "patched_local",
        "output": str(output),
        "patch_count": len(ready),
        "verify_errors": 0,
    }

    if not args.upload:
        print(json.dumps(report, ensure_ascii=False, indent=2))
        print("\n本地补丁已完成，未上传。加 --upload 执行云文档写回。", flush=True)
        return 0

    sid, sid_src = resolve_wps_sid(cfg)
    if not sid:
        print("缺少 WPS_SID", file=sys.stderr)
        return 4

    wps_root = find_wps365_read_root(cfg)
    if not wps_root:
        print("未找到 wps365-read", file=sys.stderr)
        return 4

    link_id = args.link_id or (cfg.get("dept_report") or {}).get("link_id")
    if not link_id:
        print("config 中缺少 dept_report.link_id", file=sys.stderr)
        return 4

    proc = run_drive(wps_root, ["update", link_id, str(output), "--confirm"])
    stdout = (proc.stdout or "") + (proc.stderr or "")
    report["upload"] = {"returncode": proc.returncode, "output": stdout[:2000]}
    print(stdout)
    if proc.returncode != 0:
        return 5

    report["status"] = "uploaded"
    print(json.dumps(report, ensure_ascii=False, indent=2))

    if args.keep_cache:
        (CACHE / "apply-report.json").write_text(
            json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print("\n已保留 .cache（--keep-cache）。", flush=True)
        return 0

    cleanup_cmd = [sys.executable, str(WORKFLOW / "cleanup_cache.py")]
    print(f">>> {' '.join(cleanup_cmd)}", flush=True)
    cleanup_rc = subprocess.call(cleanup_cmd, cwd=str(SKILL_ROOT))
    if cleanup_rc != 0:
        print("警告：缓存清理未完全成功，可手动检查 .cache/", file=sys.stderr)
        return 6

    print("\n已清理 .cache 中间产物。", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
