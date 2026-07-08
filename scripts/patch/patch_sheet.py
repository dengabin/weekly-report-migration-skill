#!/usr/bin/env python3
"""对 xlsx/ksheet 执行单元格级补丁（不改结构，只写值）。"""
from __future__ import annotations

import argparse
import json
import shutil
import sys
import tempfile
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as e:
    raise SystemExit("请安装 openpyxl: pip install openpyxl") from e


def load_json(path: Path) -> dict:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def cell_key(sheet: str, row: int, col: int) -> tuple[str, int, int]:
    return (sheet, row, col)


def snapshot_workbook(wb) -> dict[tuple[str, int, int], object]:
    snap: dict[tuple[str, int, int], object] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1, min_col=1, max_col=ws.max_column or 1):
            for cell in row:
                snap[cell_key(sheet_name, cell.row, cell.column)] = cell.value
    return snap


def values_equal(a, b) -> bool:
    if a == b:
        return True
    return str(a or "") == str(b or "")


def resolve_workbook_path(input_path: Path) -> tuple[Path, Path | None]:
    """ksheet 需复制为 xlsx 供 openpyxl 打开。返回 (work_path, temp_dir)。"""
    if input_path.suffix.lower() == ".ksheet":
        tmp = Path(tempfile.mkdtemp(prefix="report-migration-"))
        xlsx_path = tmp / "work.xlsx"
        shutil.copy2(input_path, xlsx_path)
        return xlsx_path, tmp
    return input_path, None


def content_by_name(extracted: dict) -> dict[str, str]:
    """写入 content（已做 •/◦ 符号转换）；content_raw 保留 otl 原文备查。"""
    return {m["name"]: m.get("content", "") for m in extracted.get("members", [])}


def set_plain_text_cell(cell, text: str, template_cell=None) -> None:
    """纯文本写入，清除超链接；复制历史列的换行/对齐/字体（微软雅黑 10）。"""
    from copy import copy

    cell.value = text
    cell.hyperlink = None

    if template_cell is not None:
        if template_cell.alignment:
            align = copy(template_cell.alignment)
            align.wrap_text = True
            align.horizontal = align.horizontal or "left"
            align.vertical = align.vertical or "center"
            cell.alignment = align
        if template_cell.font:
            src = template_cell.font
            cell.font = copy(src)
        else:
            from openpyxl.styles import Font

            cell.font = Font(name="微软雅黑", size=10)
    else:
        from openpyxl.styles import Alignment, Font

        cell.font = Font(name="微软雅黑", size=10)
        align = copy(cell.alignment) if cell.alignment else Alignment()
        align.wrap_text = True
        align.horizontal = "left"
        align.vertical = "center"
        cell.alignment = align


def apply_patches(
    wb,
    plan: dict,
    extracted: dict,
    *,
    strict_verbatim: bool = True,
) -> tuple[list[dict], list[dict]]:
    by_name = content_by_name(extracted)
    default_sheet = wb.sheetnames[0]
    applied: list[dict] = []
    skipped: list[dict] = []

    for patch in plan.get("patches", []):
        if patch.get("status") != "ready":
            skipped.append({"name": patch.get("name"), "reason": patch.get("status")})
            continue

        name = patch["name"]
        source = by_name.get(name)
        if source is None:
            skipped.append({"name": name, "reason": "extracted.json 中无该成员"})
            continue
        if strict_verbatim and patch.get("content") != source:
            skipped.append({"name": name, "reason": "patch-plan 与 extracted 内容不一致，拒绝写入"})
            continue

        sheet_name = patch.get("sheet") or patch.get("target", {}).get("sheet") or default_sheet
        if sheet_name not in wb.sheetnames:
            skipped.append({"name": name, "reason": f"sheet 不存在: {sheet_name}"})
            continue

        ws = wb[sheet_name]
        row = int(patch["row"])
        col = int(patch["col"])
        old = ws.cell(row=row, column=col).value
        target = ws.cell(row=row, column=col)
        # 参考同行后续列（跳过链接和空列）的字体与换行
        template = None
        for offset in range(1, 5):
            tc = col + offset
            if tc > (ws.max_column or 1):
                break
            c = ws.cell(row=row, column=tc)
            if c.value and str(c.value).strip() and not str(c.value).strip().startswith("📄"):
                template = c
                break
        set_plain_text_cell(target, source, template)
        applied.append(
            {
                "name": name,
                "cell": patch.get("cell"),
                "sheet": sheet_name,
                "row": row,
                "col": col,
                "old_len": len(str(old or "")),
                "new_len": len(source),
            }
        )
    return applied, skipped


def verify_non_target_unchanged(
    before: dict[tuple[str, int, int], object],
    after_wb,
    target_keys: set[tuple[str, int, int]],
) -> list[dict]:
    violations: list[dict] = []
    for sheet_name in after_wb.sheetnames:
        ws = after_wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1, min_col=1, max_col=ws.max_column or 1):
            for cell in row:
                key = cell_key(sheet_name, cell.row, cell.column)
                if key in target_keys:
                    continue
                old = before.get(key)
                new = cell.value
                if not values_equal(old, new):
                    violations.append(
                        {
                            "sheet": sheet_name,
                            "row": cell.row,
                            "col": cell.column,
                            "before": str(old)[:120] if old is not None else None,
                            "after": str(new)[:120] if new is not None else None,
                        }
                    )
    return violations


def save_workbook(wb, output: Path) -> Path:
    output.parent.mkdir(parents=True, exist_ok=True)
    if output.suffix.lower() == ".ksheet":
        tmp_xlsx = output.with_suffix(".xlsx")
        wb.save(tmp_xlsx)
        shutil.copy2(tmp_xlsx, output)
        return output
    wb.save(output)
    return output


def main() -> int:
    parser = argparse.ArgumentParser(description="补丁写入部门 xlsx/et（仅改计划内单元格）")
    parser.add_argument("--config", type=Path, default=Path("config.json"))
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--plan", type=Path, required=True)
    parser.add_argument("--extracted", type=Path, required=True, help="otl 提取结果，内容原样写入")
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--no-strict-verbatim", action="store_true", help="不校验 patch-plan 与 extracted 一致")
    args = parser.parse_args()

    if args.input.suffix.lower() == ".ksheet":
        print(
            "禁止对 .ksheet 使用 openpyxl 保存（会破坏 customXml 超链接）。"
            "请改用 scripts/patch/patch_ksheet_zip.py",
            file=sys.stderr,
        )
        return 1

    cfg = load_json(args.config)
    plan = load_json(args.plan)
    extracted = load_json(args.extracted)

    open_path, temp_dir = resolve_workbook_path(args.input)
    try:
        wb = load_workbook(open_path)
        before = snapshot_workbook(wb)

        applied, skipped = apply_patches(
            wb,
            plan,
            extracted,
            strict_verbatim=not args.no_strict_verbatim,
        )
        if skipped:
            print(json.dumps({"applied": applied, "skipped": skipped}, ensure_ascii=False, indent=2), file=sys.stderr)
            return 1

        target_keys = {cell_key(p["sheet"], p["row"], p["col"]) for p in applied}
        violations = verify_non_target_unchanged(before, wb, target_keys)
        if violations:
            print(
                json.dumps({"error": "非目标单元格发生变化", "violations": violations[:20]}, ensure_ascii=False, indent=2),
                file=sys.stderr,
            )
            return 2

        final = save_workbook(wb, args.output)
        summary = {
            "week": cfg.get("week"),
            "input": str(args.input),
            "output": str(final),
            "patch_count": len(applied),
            "applied": applied,
            "skipped": skipped,
            "non_target_violations": 0,
        }
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0
    finally:
        if temp_dir and temp_dir.exists():
            shutil.rmtree(temp_dir, ignore_errors=True)


if __name__ == "__main__":
    raise SystemExit(main())
